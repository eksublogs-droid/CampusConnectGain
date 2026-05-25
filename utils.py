"""
CampusConnect — Utilities
"""
import os
import io
import json
import time
import hmac
import hashlib
import requests
import tempfile
import vobject
import openpyxl
from datetime import datetime

FLW_SECRET = os.getenv("FLW_SECRET_KEY", "")
FLW_BASE = "https://api.flutterwave.com/v3"

NIGERIAN_STATES = [
    "Abia","Adamawa","Akwa Ibom","Anambra","Bauchi","Bayelsa","Benue","Borno",
    "Cross River","Delta","Ebonyi","Edo","Ekiti","Enugu","FCT","Gombe","Imo",
    "Jigawa","Kaduna","Kano","Katsina","Kebbi","Kogi","Kwara","Lagos","Nasarawa",
    "Niger","Ogun","Ondo","Osun","Oyo","Plateau","Rivers","Sokoto","Taraba",
    "Yobe","Zamfara"
]

LEVELS = ["100L", "200L", "300L", "400L", "500L", "600L", "Postgraduate", "Alumni"]

INTERESTS = [
    "Tech & Coding","Business & Entrepreneurship","Graphics & Design",
    "Music & Arts","Sports & Fitness","Fashion & Style","Media & Content",
    "Finance & Investment","Law & Politics","Health & Medicine",
    "Education & Teaching","Agriculture","Engineering"
]

AD_TIERS = {
    "basic": {"price": 500, "label": "🔹 Basic", "duration_hours": 24, "features": "Text post · 24hrs pinned · ~500-2k impressions"},
    "standard": {"price": 1500, "label": "🔶 Standard", "duration_hours": 48, "features": "Text + image · 48hrs pinned · Broadcast to all users · ~2k-8k impressions"},
    "premium": {"price": 4000, "label": "💎 Premium", "duration_hours": 168, "features": "Text + image/video · 7 days pinned · Targeted DMs · Analytics report"},
}


# ────────────────────────────────────────
# PROFILE CARD FORMATTING
# ────────────────────────────────────────

def format_profile_card(user, viewer_id=None, is_own_profile=False):
    interests_str = ", ".join(user.get('interests') or []) or "—"
    bio = user.get('bio') or "—"

    whatsapp = user.get('whatsapp_number') or user.get('phone')
    show_wa = user.get('show_whatsapp', True)
    owner_viewing = is_own_profile or (viewer_id is not None and viewer_id == user.get('id'))

    if owner_viewing:
        wa_line = f"💬 WhatsApp: {whatsapp}"
    elif show_wa:
        wa_line = f"💬 WhatsApp: {whatsapp}"
    else:
        wa_line = "💬 WhatsApp: _Contact on request_"

    return (
        f"👤 *{user['full_name']}*\n"
        f"🎓 {user['school']} · {user['department']} · {user['level']}\n"
        f"📍 {user['state']}\n"
        f"📱 {user['phone']}\n"
        f"{wa_line}\n"
        f"🌟 Interests: {interests_str}\n"
        f"💬 {bio}"
    )


def format_profile_summary(user):
    return (
        f"*{user['full_name']}*\n"
        f"🎓 {user['school']} · {user['level']}\n"
        f"📍 {user['state']} · {user['department']}"
    )


def format_ad_channel_post(ad, user):
    tier_emoji = {"basic": "🔹", "standard": "🔶", "premium": "💎"}.get(ad['tier'], "📢")
    return (
        f"📢 *SPONSORED* {tier_emoji}\n\n"
        f"{ad['ad_copy']}\n\n"
        f"👤 {user['full_name']} · {user['school']}\n"
        f"📱 {user['phone']}\n"
        f"━━━━━━━━━━━━━━\n"
        f"_CampusConnect · Nigeria's Student Network_"
    )


def format_reg_channel_post(user):
    interests_str = ", ".join(user.get('interests') or []) or "—"
    return (
        f"🎉 *New Student Joined CampusConnect!*\n\n"
        f"👤 *{user['full_name']}*\n"
        f"🎓 {user['school']} · {user['department']} · {user['level']}\n"
        f"📍 {user['state']}\n"
        f"🌟 {interests_str}\n\n"
        f"_Connect with thousands of Nigerian students →_ @CampusConnectBot"
    )


# ────────────────────────────────────────
# VCF GENERATION
# ────────────────────────────────────────

def generate_vcf(users: list) -> bytes:
    output = io.StringIO()
    for u in users:
        vcard = vobject.vCard()
        vcard.add('n').value = vobject.vcard.Name(family=u['full_name'].split()[-1] if ' ' in u['full_name'] else u['full_name'], given=u['full_name'].split()[0])
        vcard.add('fn').value = u['full_name']
        tel = vcard.add('tel')
        tel.value = u['phone']
        tel.type_param = 'CELL'
        org = vcard.add('org')
        org.value = [u['school'], u['department']]
        note = vcard.add('note')
        note.value = f"{u['level']} · {u['state']} · CampusConnect"
        output.write(vcard.serialize())
    return output.getvalue().encode('utf-8')


# ────────────────────────────────────────
# EXCEL EXPORT
# ────────────────────────────────────────

def generate_excel(users: list) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CampusConnect Students"

    headers = [
        "ID", "Full Name", "Username", "Phone", "WhatsApp", "Date of Birth",
        "Show WhatsApp", "School", "Department", "Level", "State",
        "Interests", "Bio", "Registered At"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True)

    for row_idx, user in enumerate(users, 2):
        ws.cell(row=row_idx, column=1, value=str(user['id']))
        ws.cell(row=row_idx, column=2, value=user['full_name'])
        ws.cell(row=row_idx, column=3, value=user.get('username') or "")
        ws.cell(row=row_idx, column=4, value=user['phone'])
        ws.cell(row=row_idx, column=5, value=user.get('whatsapp_number') or "")
        ws.cell(row=row_idx, column=6, value=str(user.get('date_of_birth') or ""))
        ws.cell(row=row_idx, column=7, value="Yes" if user.get('show_whatsapp', True) else "No")
        ws.cell(row=row_idx, column=8, value=user['school'])
        ws.cell(row=row_idx, column=9, value=user['department'])
        ws.cell(row=row_idx, column=10, value=user['level'])
        ws.cell(row=row_idx, column=11, value=user['state'])
        ws.cell(row=row_idx, column=12, value=", ".join(user.get('interests') or []))
        ws.cell(row=row_idx, column=13, value=user.get('bio') or "")
        ws.cell(row=row_idx, column=14, value=str(user['registered_at']))

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ────────────────────────────────────────
# FLUTTERWAVE — VIRTUAL ACCOUNT
# ────────────────────────────────────────

def create_virtual_account(tx_ref: str, amount: int, narration: str, meta: dict = None) -> dict:
    """
    Create a temporary virtual account via Flutterwave.
    Returns dict with: account_number, bank_name, tx_ref, expires_at
    """
    headers = {
        "Authorization": f"Bearer {FLW_SECRET}",
        "Content-Type": "application/json"
    }
    payload = {
        "email": f"pay+{tx_ref}@campusconnect.ng",
        "amount": amount,
        "currency": "NGN",
        "tx_ref": tx_ref,
        "narration": narration,
        "is_permanent": False,
        "meta": meta or {}
    }
    r = requests.post(f"{FLW_BASE}/virtual-account-numbers", headers=headers, json=payload, timeout=15)
    data = r.json()
    if data.get("status") == "success":
        d = data["data"]
        return {
            "account_number": d["account_number"],
            "bank_name": d["bank_name"],
            "tx_ref": tx_ref,
            "flw_ref": d.get("flw_ref", ""),
            "amount": amount,
        }
    raise Exception(f"Flutterwave error: {data.get('message', 'Unknown error')}")


def verify_flw_payment(tx_ref: str) -> tuple[bool, dict]:
    """Verify a Flutterwave payment by tx_ref."""
    headers = {"Authorization": f"Bearer {FLW_SECRET}"}
    r = requests.get(
        f"{FLW_BASE}/transactions",
        headers=headers,
        params={"tx_ref": tx_ref},
        timeout=15
    )
    data = r.json()
    if data.get("status") == "success":
        txns = data.get("data", [])
        for txn in txns:
            if txn.get("status") == "successful":
                return True, txn
    return False, {}


def verify_flw_signature(payload: bytes, signature: str) -> bool:
    """Verify Flutterwave webhook signature."""
    secret_hash = os.getenv("FLW_WEBHOOK_SECRET", "")
    if not secret_hash:
        return True  # skip if not configured
    expected = hmac.new(secret_hash.encode(), payload, hashlib.sha256).hexdigest()
    return hmac.compare_digest(expected, signature or "")


def generate_reference(prefix: str, user_id: int) -> str:
    return f"{prefix}_{user_id}_{int(time.time())}"


def format_payment_message(acct: dict, description: str) -> str:
    """Format the bank transfer instruction message sent to user."""
    return (
        f"🏦 *Pay via Bank Transfer*\n\n"
        f"📋 {description}\n\n"
        f"*Amount:* ₦{acct['amount']:,}\n"
        f"*Bank:* {acct['bank_name']}\n"
        f"*Account Number:* `{acct['account_number']}`\n\n"
        f"⏱ Transfer exactly ₦{acct['amount']:,} to the account above.\n"
        f"✅ Payment is confirmed automatically once received.\n\n"
        f"_Use the button below to check if payment was received._"
    )


# ────────────────────────────────────────
# GOOGLE SHEETS SYNC
# ────────────────────────────────────────

def sync_to_google_sheets(users: list):
    try:
        import gspread
        from google.oauth2.service_account import Credentials

        creds_json = os.getenv("GOOGLE_CREDENTIALS_JSON")
        sheet_id = os.getenv("GOOGLE_SHEETS_ID")
        if not creds_json or not sheet_id:
            return False

        creds = Credentials.from_service_account_info(
            json.loads(creds_json),
            scopes=["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
        )
        client = gspread.authorize(creds)
        sheet = client.open_by_key(sheet_id).sheet1

        sheet.clear()
        headers = ["ID", "Name", "Username", "Phone", "School", "Dept", "Level", "State", "Interests", "Registered"]
        sheet.append_row(headers)
        for u in users:
            sheet.append_row([
                str(u['id']), u['full_name'], u.get('username') or "",
                u['phone'], u['school'], u['department'],
                u['level'], u['state'],
                ", ".join(u.get('interests') or []),
                str(u['registered_at'])
            ])
        return True
    except Exception as e:
        print(f"Sheets sync error: {e}")
        return False


# ────────────────────────────────────────
# ANTI-SPAM
# ────────────────────────────────────────

_rate_limit = {}

def check_rate_limit(user_id: int, action: str, max_per_minute: int = 10) -> bool:
    key = f"{user_id}:{action}"
    now = time.time()
    window = _rate_limit.get(key, [])
    window = [t for t in window if now - t < 60]
    if len(window) >= max_per_minute:
        return False
    window.append(now)
    _rate_limit[key] = window
    return True


def validate_phone(phone: str):
    digits = ''.join(c for c in phone if c.isdigit())
    if len(digits) == 11 and digits.startswith('0'):
        return '+234' + digits[1:]
    elif len(digits) == 13 and digits.startswith('234'):
        return '+' + digits
    elif len(digits) == 10:
        return '+234' + digits
    return None


def format_whatsapp(number: str) -> str | None:
    return validate_phone(number)


def format_birthday_message(user: dict) -> str:
    name = user['full_name'].split()[0]
    return (
        f"🎂 *Happy Birthday, {name}!* 🎉\n\n"
        f"CampusConnect wishes you an amazing day filled with joy!\n\n"
        f"_Keep connecting, keep growing. 🚀_"
    )


# ────────────────────────────────────────
# KEYBOARDS
# ────────────────────────────────────────

from telegram import InlineKeyboardButton, InlineKeyboardMarkup, ReplyKeyboardMarkup, KeyboardButton


def main_menu_keyboard():
    return ReplyKeyboardMarkup([
        ["📋 My Profile", "✏️ Edit Profile"],
        ["🔍 Find Students", "📚 Directory"],
        ["🛒 Store", "📬 Contact Drop"],
        ["📢 Run Ad", "🔖 Saved"],
        ["ℹ️ Help"]
    ], resize_keyboard=True)


def cancel_keyboard():
    return ReplyKeyboardMarkup([["❌ Cancel"]], resize_keyboard=True, one_time_keyboard=True)


def states_keyboard():
    rows = []
    state_list = NIGERIAN_STATES
    for i in range(0, len(state_list), 3):
        rows.append(state_list[i:i+3])
    rows.append(["❌ Cancel"])
    return ReplyKeyboardMarkup(rows, resize_keyboard=True, one_time_keyboard=True)


def levels_keyboard():
    return ReplyKeyboardMarkup(
        [LEVELS[:4], LEVELS[4:], ["❌ Cancel"]],
        resize_keyboard=True, one_time_keyboard=True
    )


def interests_keyboard():
    rows = []
    for i in range(0, len(INTERESTS), 2):
        rows.append(INTERESTS[i:i+2])
    rows.append(["✅ Done", "❌ Cancel"])
    return ReplyKeyboardMarkup(rows, resize_keyboard=True)


def profile_action_keyboard(target_id: int):
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("📲 Connect", callback_data=f"connect:{target_id}"),
         InlineKeyboardButton("🔖 Save", callback_data=f"save:{target_id}")],
        [InlineKeyboardButton("🚩 Report", callback_data=f"report:{target_id}")]
    ])


def pagination_keyboard(page: int, total_pages: int, prefix: str):
    buttons = []
    if page > 0:
        buttons.append(InlineKeyboardButton("◀ Prev", callback_data=f"{prefix}:page:{page-1}"))
    if page < total_pages - 1:
        buttons.append(InlineKeyboardButton("Next ▶", callback_data=f"{prefix}:page:{page+1}"))
    return InlineKeyboardMarkup([buttons]) if buttons else None


def ad_tier_keyboard():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("🔹 Basic — ₦500", callback_data="adtier:basic")],
        [InlineKeyboardButton("🔶 Standard — ₦1,500", callback_data="adtier:standard")],
        [InlineKeyboardButton("💎 Premium — ₦4,000", callback_data="adtier:premium")],
    ])


def store_category_keyboard():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("📇 Contact Packs", callback_data="store:contact_pack")],
        [InlineKeyboardButton("🔗 Group Links", callback_data="store:group_link")],
        [InlineKeyboardButton("🛠️ Tools", callback_data="store:tool")],
        [InlineKeyboardButton("🎛️ Custom Order", callback_data="store:custom")],
    ])


def drop_tier_keyboard():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("🆓 Free Drop (New students, 3 days)", callback_data="drop:free")],
        [InlineKeyboardButton("⭐ Premium Drop — ₦500 (Full DB)", callback_data="drop:premium")],
    ])
